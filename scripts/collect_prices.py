#!/usr/bin/env python3
"""
Fuel Price Collector for Lithuania (ena.lt)
Scrapes fuel prices from the Lithuanian Energy Agency and stores them in Supabase.

Usage:
    python collect_prices.py

Environment variables:
    SUPABASE_URL - Your Supabase project URL
    SUPABASE_KEY - Your Supabase service_role key (not anon key, needs insert permission)
"""

import io
import os
import re
import sys
from collections import defaultdict

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from supabase import create_client

# --- Configuration ---
ENA_URLS = [
    "https://www.ena.lt/degalu-kainos-degalinese/",    # Main page (latest date)
    "https://www.ena.lt/dk-visa-informacija",            # Archive page (all dates)
]
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")


def scrape_all_excel_urls() -> dict[str, str]:
    """Scrape ena.lt pages and extract all dated Excel file URLs."""
    date_pattern = re.compile(r"\d{4}-\d{2}-\d{2}")
    result = {}

    for page_url in ENA_URLS:
        print(f"Scraping {page_url}...")
        try:
            resp = requests.get(page_url, timeout=30)
            resp.raise_for_status()
        except Exception as e:
            print(f"  Failed to fetch {page_url}: {e}")
            continue

        soup = BeautifulSoup(resp.text, "html.parser")
        links = soup.select("a[href*='sharepoint.com']")

        for link in links:
            href = link.get("href", "")
            title = link.get("title", "")
            text = link.get_text()

            # Only match Excel links (:x: in URL) with "kain" in title or text
            if ":x:" not in href:
                continue
            combined = f"{text} {title}".lower()
            if "kain" not in combined:
                continue

            search_text = f"{text} {title} {href}"
            match = date_pattern.search(search_text)
            if match and match.group() not in result:
                result[match.group()] = href

    print(f"Found {len(result)} dated Excel files: {sorted(result.keys())}")
    return result


def download_excel(url: str) -> bytes:
    """Download Excel file from SharePoint with cookie-based auth."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })

    # Step 1: Request with download=1, don't follow redirects
    download_url = f"{url}&download=1" if "?" in url else f"{url}?download=1"
    resp = session.get(download_url, allow_redirects=False, timeout=30)

    if resp.status_code in (301, 302, 303, 307, 308):
        # Step 2: Follow redirect with cookies (FedAuth)
        redirect_url = resp.headers.get("Location", "")
        if redirect_url and not redirect_url.startswith("http"):
            from urllib.parse import urlparse
            parsed = urlparse(download_url)
            redirect_url = f"{parsed.scheme}://{parsed.netloc}{redirect_url}"

        resp2 = session.get(redirect_url, timeout=60)
        resp2.raise_for_status()
        data = resp2.content
    else:
        resp.raise_for_status()
        data = resp.content

    if len(data) < 100:
        raise ValueError(f"Downloaded file too small ({len(data)} bytes)")

    return data


def parse_excel(data: bytes, override_date: str) -> list[dict]:
    """Parse Excel file, auto-detecting wide vs long format."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb.active

    # Find header row
    header_row = None
    header_idx = None
    for i, row in enumerate(sheet.iter_rows(max_row=16, values_only=False)):
        values = [str(cell.value).lower().strip() if cell.value else "" for cell in row]
        has_data = values[0] == "data" if values else False
        has_fuel = any("benzin" in v or "dyzel" in v or "snd" in v for v in values)
        has_company = any("mon" in v and "tinkl" in v for v in values)
        has_long = any("degalai" in v or "degalu" in v for v in values) and any("kaina" in v for v in values)

        if has_data and (has_fuel or has_company or has_long):
            header_row = values
            header_idx = i
            break

    if header_row is None:
        print(f"  WARNING: Could not find header row for {override_date}")
        wb.close()
        return []

    # Map columns
    col_map = {}
    for i, h in enumerate(header_row):
        if ("mon" in h and "tinkl" in h) or "bendrov" in h or "pavadinim" in h:
            col_map["company"] = i
        elif "gatv" in h or "gyvenviet" in h:
            col_map["address"] = i
        elif "savivaldyb" in h or "rajon" in h:
            col_map["municipality"] = i
        elif "95" in h and "98" not in h:
            col_map["petrol95"] = i
        elif "98" in h:
            col_map["petrol98"] = i
        elif "dyzel" in h or "diesel" in h:
            col_map["diesel"] = i
        elif "lpg" in h or "snd" in h or "duj" in h or "suskystint" in h:
            col_map["lpg"] = i
        elif "degalai" in h or h == "degalu":
            col_map["fuel_name"] = i
        elif "kaina" in h:
            col_map["price"] = i

    is_long = "fuel_name" in col_map and "price" in col_map

    if is_long:
        prices = parse_long_format(sheet, header_idx, col_map, override_date)
    else:
        prices = parse_wide_format(sheet, header_idx, col_map, override_date)

    wb.close()
    return prices


def parse_wide_format(sheet, header_idx: int, col_map: dict, date: str) -> list[dict]:
    """One row per station, fuel types as columns."""
    prices = []
    company_col = col_map.get("company")
    if company_col is None:
        return prices

    for row in sheet.iter_rows(min_row=header_idx + 2, values_only=True):
        company = safe_str(row, company_col)
        if not company:
            continue

        prices.append({
            "date": date,
            "company": company.strip(),
            "address": safe_str(row, col_map.get("address")).strip(),
            "municipality": safe_str(row, col_map.get("municipality")).strip(),
            "petrol95": safe_float(row, col_map.get("petrol95")),
            "petrol98": safe_float(row, col_map.get("petrol98")),
            "diesel": safe_float(row, col_map.get("diesel")),
            "lpg": safe_float(row, col_map.get("lpg")),
        })

    return prices


def parse_long_format(sheet, header_idx: int, col_map: dict, date: str) -> list[dict]:
    """One row per fuel type per station — pivot to wide."""
    company_col = col_map.get("company")
    fuel_col = col_map.get("fuel_name")
    price_col = col_map.get("price")
    if company_col is None or fuel_col is None or price_col is None:
        return []

    stations = defaultdict(lambda: {
        "date": date, "company": "", "address": "", "municipality": "",
        "petrol95": None, "petrol98": None, "diesel": None, "lpg": None
    })

    for row in sheet.iter_rows(min_row=header_idx + 2, values_only=True):
        company = safe_str(row, company_col).strip()
        if not company:
            continue

        address = safe_str(row, col_map.get("address")).strip()
        municipality = safe_str(row, col_map.get("municipality")).strip()
        fuel_name = safe_str(row, fuel_col).lower().strip()
        price = safe_float(row, price_col)

        if price is None:
            continue

        key = (company, address)
        entry = stations[key]
        entry["company"] = company
        entry["address"] = address
        entry["municipality"] = municipality

        if "95" in fuel_name:
            entry["petrol95"] = price
        elif "98" in fuel_name:
            entry["petrol98"] = price
        elif "dyzel" in fuel_name or "diesel" in fuel_name:
            entry["diesel"] = price
        elif "suskystint" in fuel_name or "snd" in fuel_name or "lpg" in fuel_name or "duj" in fuel_name:
            entry["lpg"] = price

    return list(stations.values())


def safe_str(row: tuple, col: int | None) -> str:
    if col is None or col >= len(row) or row[col] is None:
        return ""
    return str(row[col])


def safe_float(row: tuple, col: int | None) -> float | None:
    if col is None or col >= len(row) or row[col] is None:
        return None
    val = row[col]
    if isinstance(val, (int, float)):
        return float(val) if val > 0 else None
    s = str(val).strip().replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def geocode_prices(supabase_client, prices: list[dict]):
    """Add lat/lng to prices using cached geocode data or Nominatim."""
    import time
    from geopy.geocoders import Nominatim
    from geopy.exc import GeocoderTimedOut

    # Load cached locations from Supabase
    cached = {}
    try:
        result = supabase_client.table("station_locations").select("*").execute()
        for row in (result.data or []):
            cached[(row["company"], row["address"])] = (row["latitude"], row["longitude"])
    except Exception:
        pass  # Table might not exist yet

    geolocator = Nominatim(user_agent="fuel_prices_lt", timeout=10)
    new_geocodes = []

    for price in prices:
        key = (price["company"], price["address"])
        if key in cached:
            price["latitude"] = cached[key][0]
            price["longitude"] = cached[key][1]
            continue

        # Try geocoding
        try:
            query = f"{price['address']}, {price.get('municipality', '')}, Lithuania"
            location = geolocator.geocode(query)
            if location:
                price["latitude"] = round(location.latitude, 6)
                price["longitude"] = round(location.longitude, 6)
                cached[key] = (price["latitude"], price["longitude"])
                new_geocodes.append({
                    "company": price["company"],
                    "address": price["address"],
                    "latitude": price["latitude"],
                    "longitude": price["longitude"]
                })
            else:
                price["latitude"] = None
                price["longitude"] = None
            time.sleep(1.1)  # Nominatim rate limit: 1 req/sec
        except (GeocoderTimedOut, Exception) as e:
            print(f"    Geocode failed for {price['address']}: {e}")
            price["latitude"] = None
            price["longitude"] = None

    # Cache new geocodes
    if new_geocodes:
        try:
            supabase_client.table("station_locations").upsert(
                new_geocodes, on_conflict="company,address"
            ).execute()
            print(f"  Cached {len(new_geocodes)} new geocoded locations")
        except Exception as e:
            print(f"  Failed to cache geocodes: {e}")

    geocoded = sum(1 for p in prices if p.get("latitude"))
    print(f"  Geocoded: {geocoded}/{len(prices)} stations")


def upsert_to_supabase(supabase_client, prices: list[dict], date: str):
    """Upsert prices to Supabase with conflict handling."""
    if not prices:
        print(f"  No prices to upsert for {date}")
        return

    # Supabase upsert with ON CONFLICT (date, company, address)
    batch_size = 500
    total = 0
    for i in range(0, len(prices), batch_size):
        batch = prices[i:i + batch_size]
        supabase_client.table("fuel_prices").upsert(
            batch,
            on_conflict="date,company,address"
        ).execute()
        total += len(batch)

    print(f"  Upserted {total} records for {date}")


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Set SUPABASE_URL and SUPABASE_KEY environment variables")
        sys.exit(1)

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Get existing dates in Supabase
    existing = supabase.table("fuel_prices").select("date").execute()
    existing_dates = set(row["date"] for row in existing.data) if existing.data else set()
    print(f"Existing dates in Supabase: {sorted(existing_dates)}")

    # Scrape all available Excel URLs from ena.lt
    urls = scrape_all_excel_urls()

    for date, url in sorted(urls.items()):
        if date in existing_dates:
            print(f"Skipping {date} — already in Supabase")
            continue

        print(f"Processing {date}...")
        try:
            excel_data = download_excel(url)
            print(f"  Downloaded {len(excel_data)} bytes")

            prices = parse_excel(excel_data, date)
            print(f"  Parsed {len(prices)} station records")

            geocode_prices(supabase, prices)
            upsert_to_supabase(supabase, prices, date)
        except Exception as e:
            print(f"  ERROR processing {date}: {e}")
            continue

    print("Done!")


if __name__ == "__main__":
    main()
